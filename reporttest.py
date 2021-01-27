import linecache
from openpyxl import Workbook
from openpyxl.styles import Font, Fill
from openpyxl.styles import Alignment
import re
import os
from lxml import etree
import arrow


directory = "yourdir"

p = re.compile('>(.*)<')


for file in os.listdir(directory):
    listel = []
    wb = Workbook()
    ws = wb.active
    dest_filename = 'C:/tmp/excel/book_' + file + '_.xlsx'

    ws['A1'] = 'serial number'
    ws.column_dimensions['A'].width = 15
    ws['A1'].font = Font(bold=True)

    ws['B1'] = 'IR number'
    ws.column_dimensions['B'].width = 14
    ws['B1'].font = Font(bold=True)

    ws['C1'] = 'mark'
    ws.column_dimensions['C'].width = 21
    ws['C1'].font = Font(bold=True)

    ws['D1'] = 'status date'
    ws.column_dimensions['D'].width = 12
    ws['D1'].font = Font(bold=True)

    ws['E1'] = 'classes'
    ws.column_dimensions['E'].width = 9
    ws['E1'].font = Font(bold=True)

    ws['F1'] = 'owner'
    ws.column_dimensions['F'].width = 51
    ws['F1'].font = Font(bold=True)

    ws['G1'] = 'USPTO deadline'
    ws.column_dimensions['G'].width = 15
    ws['G1'].font = Font(bold=True)

    ws['H1'] = 'correspondent'
    ws.column_dimensions['H'].width = 43
    ws['H1'].font = Font(bold=True)

    i = 2
    context = etree.iterparse(
        directory + "/" + file,
        events=(
            'end',
        ),
        tag='case-file')

    for event, element in context:
        for item in element.iter():

            if item.tag == 'attorney-name':
                element.clear()
                break

            elif item.tag == 'state':
                element.clear()
                break

            elif item.tag == 'country' and item.text == 'US':
                element.clear()
                break

            elif item.tag == 'country' and 'U.S.' in item.text:
                element.clear()
                break

        else:

            for item in element.iter():

                if item.tag == 'status-code' and item.text == '641' and element[3][0].tag == 'filing-date':

                    listel.append(element)

    for elem in listel:
        j = 0
        p = 0
        c = 0
        k = 0
        kontrol = 0
        for item in elem.iter():
            if item.tag == 'serial-number':
                print("serial " + item.text)
                ws['A' + str(i)] = str(item.text)
            elif item.tag == 'status-date' and kontrol == 0:
                print(item.text + " tag " + item.tag)
                ws['D' + str(i)] = item.text
                date = item.text
                date1 = arrow.get(date)
                date3 = date1.shift(months=6).date()
                ws['G' + str(i)] = str(date3)
                kontrol = 1
            elif item.tag == 'international-registration-number':
                ws['B' + str(i)] = item.text
            elif item.tag == 'mark-identification':
                ws['C' + str(i)] = item.text
            elif item.tag == 'international-code':
                ws['E' + str(i + j)] = item.text
                j += 1

            elif item.tag == 'correspondent':
                for item1 in item.iter():
                    if 'address' in item1.tag:

                        text = item1.text
                        text = text.strip()

                        if text[-1] == ',':

                            text = text[:-1]
                        if text[-1] == ',':
                            text = text[:-1]

                        ws['H' + str(i + k)] = text
                        k += 1
            elif item.tag == 'case-file-owner':
                k1 = 0
                for item1 in item.iter():

                    if item1.tag == 'party-name':
                        ws['F' + str(i + p)] = item1.text
                        p += 1
                        k1 = 1
                    elif 'address' in item1.tag:
                        print('case owner address' + item1.text)
                        text = item1.text
                        text = text.strip()

                        if text[-1] == ',':

                            text = text[:-1]
                        if text[-1] == ',':
                            text = text[:-1]

                        ws['F' + str(i + p)] = text
                        p += 1
                    elif 'city' in item1.tag:
                        print('case owner address' + item1.text)
                        ws['F' + str(i + p)] = item1.text
                        p += 1

                    elif 'country' in item1.tag and k1 == 1:
                        print('case owner address' + item1.text)
                        ws['F' + str(i + p)] = item1.text
                        p += 1
                    elif 'postcode' in item1.tag:
                        print('case owner address' + item1.text)
                        ws['F' + str(i + p)] = item1.text
                        p += 1

        listmax = [j, k, c, p]

        i = i + 1 + max(listmax)

    for cell in ws['A']:
        cell.alignment = Alignment(horizontal='center')
    for cell in ws['B']:
        cell.alignment = Alignment(horizontal='center')
    for cell in ws['C']:
        cell.alignment = Alignment(horizontal='center')
    for cell in ws['D']:
        cell.alignment = Alignment(horizontal='center')
    for cell in ws['E']:
        cell.alignment = Alignment(horizontal='center')
    for cell in ws['F']:
        cell.alignment = Alignment(horizontal='center')
    for cell in ws['G']:
        cell.alignment = Alignment(horizontal='center')
    for cell in ws['H']:
        cell.alignment = Alignment(horizontal='center')

    wb.save(filename=dest_filename)
    print(file + " processed")
